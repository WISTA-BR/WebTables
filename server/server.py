import os
from datetime import datetime
from flask import Flask, request, jsonify
from flask_cors import CORS
from openpyxl import Workbook, load_workbook
from bs4 import BeautifulSoup
import gnupg

app = Flask(__name__)

# Allowed origins (update with your actual extension ID and allowed domains)
ALLOWED_ORIGINS = [
    "chrome-extension://your_extension_id",  # Replace with your actual extension id.
    "http://127.0.0.1"
]
CORS(app, resources={r"/copy_table": {"origins": ALLOWED_ORIGINS}})

# Expected authentication token.
EXPECTED_TOKEN = "my_secret_token"

# Path to Excel workbook.
EXCEL_FILE = "Web_Tables.xlsx"

# Setup GnuPG instance (ensure GPG is installed and configured on the server).
gpg = gnupg.GPG(gnupghome="gpg_keys")  # Create/use a directory (e.g., "gpg_keys") for keys.

# Load server private key fingerprint (assume you already imported/generated the key).
SERVER_PRIVATE_FINGERPRINT = "YOUR_PRIVATE_KEY_FINGERPRINT"  # Replace with your key fingerprint.

def get_workbook(filename=EXCEL_FILE):
    if os.path.exists(filename):
        wb = load_workbook(filename)
    else:
        wb = Workbook()
        default_sheet = wb.active
        wb.remove(default_sheet)
    return wb

def sanitize_title(title):
    # Remove disallowed characters for Excel sheet names: []:*?/\\
    invalid_chars = '[]:*?/\\'
    for char in invalid_chars:
        title = title.replace(char, '')
    return title.strip()[:31] if title.strip() else "Untitled"

def table_to_grid(table_html):
    """
    Use BeautifulSoup to convert table HTML into a grid (list of rows),
    each row being a list of cell texts.
    """
    grid = []
    soup = BeautifulSoup(table_html, "html.parser")
    rows = soup.find_all("tr")
    for tr in rows:
        row = []
        for cell in tr.find_all(["th", "td"]):
            row.append(cell.get_text(strip=True))
        if row:
            grid.append(row)
    return grid

def update_exists(ws):
    today = datetime.now().date().isoformat()
    for row in ws.iter_rows(min_row=1, values_only=True):
        if row and isinstance(row[0], str) and row[0].startswith("Update Timestamp:"):
            ts = row[0].replace("Update Timestamp:", "").strip()
            if ts.startswith(today):
                return True
    return False

def save_update(title, timestamp, table_html_list):
    wb = get_workbook()
    safe_title = sanitize_title(title)
    ws = wb[safe_title] if safe_title in wb.sheetnames else wb.create_sheet(title=safe_title)
    
    if update_exists(ws):
        print("Update for today already exists; skipping update.")
        wb.save(EXCEL_FILE)
        return False

    ws.append([f"Update Timestamp: {timestamp}"])
    for idx, table_html in enumerate(table_html_list):
        grid = table_to_grid(table_html)
        if not grid:
            continue
        ws.append([f"Table {idx + 1}"])
        for grid_row in grid:
            ws.append(grid_row)
        ws.append([])
    wb.save(EXCEL_FILE)
    return True

def decrypt_message(encrypted_data):
    """
    Decrypt the encrypted payload using the server's private key.
    Returns the decrypted text if successful.
    """
    decrypted = gpg.decrypt(encrypted_data, passphrase="your_passphrase_here")
    if not decrypted.ok:
        raise ValueError("Decryption failed: " + decrypted.status)
    return str(decrypted)

@app.before_request
def verify_request():
    # Check authentication token.
    token = request.headers.get("X-Auth-Token")
    if token != EXPECTED_TOKEN:
        return jsonify({"status": "error", "message": "Unauthorized: invalid token"}), 401
    # Refined origin checking.
    origin = request.headers.get("Origin")
    if origin and origin not in ALLOWED_ORIGINS:
        return jsonify({"status": "error", "message": "Origin not allowed"}), 403

@app.route('/copy_table', methods=['POST'])
def copy_table():
    try:
        data = request.get_json()
        if "encryptedData" not in data:
            return jsonify({"status": "error", "message": "Missing encryptedData payload"}), 400

        # Decrypt the incoming payload.
        decrypted_text = decrypt_message(data["encryptedData"])
        # The decrypted text is a JSON string.
        import json
        payload = json.loads(decrypted_text)
        
        title = payload.get("title", "Untitled_Page")
        table_html_list = payload.get("tables", [])
        timestamp = payload.get("timestamp", datetime.now().isoformat())
        
        updated = save_update(title, timestamp, table_html_list)
        message = "Data saved to Excel." if updated else "Data for today already exists; update skipped."
        return jsonify({"status": "success", "message": message})
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500

if __name__ == '__main__':
    app.run(port=5000)
